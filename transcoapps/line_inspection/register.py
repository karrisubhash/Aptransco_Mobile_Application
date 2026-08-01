"""Line inspection register — the formal patrolling record, a per-line matrix
that mirrors the physical field register (see the "Register" tab in the POCs
poc_v2.html / gisdata/clearpoc.html).

Layout: columns = the towers on one line (each showing its LATEST inspection for
the chosen cycle), rows = metadata (Date / Loc No / Tower type) + checklist items
grouped by component category + a Remarks row. Each item×tower cell is
blank / N/A / Not available / ✓ / defect(+criticality), per-position for
positional items — mirroring the POC's itemCell().

build_register() produces a plain structure reused by both the HTML view and the
Excel export (register_to_xlsx). Read-only; no capture happens here.
"""
import re
from io import BytesIO

from .models import Tower, ChecklistItemGroup, FollowUpQuestion
from . import status as tower_status
from . import schedule as sched
from .services.criticality import worst_of

CYCLE_LABELS = {
    '': 'Latest (any cycle)',
    'ground_patrol': 'Ground Patrolling',
    'pmi': 'Pre-Monsoon Inspection (PMI)',
}


def _tower_sort_key(tower):
    """Best-effort numeric sort on tower_number ('12', '12/1', 'A-3'): numeric
    first (by value), then the rest lexicographically."""
    m = re.match(r'\s*(\d+)', tower.tower_number or '')
    return (0, int(m.group(1)), tower.tower_number or '') if m else (1, 0, tower.tower_number or '')


def _answers_text(answers, fu_map):
    parts = []
    for key, value in (answers or {}).items():
        q = fu_map.get(key)
        label = q.question_text if q else key
        unit = (q.unit if q else '') or ''
        if isinstance(value, list):
            value = ', '.join(str(x) for x in value)
        parts.append(f"{label}: {value}{(' ' + unit) if unit else ''}")
    return '; '.join(parts)


def _item_cell(item, inspection, fu_map):
    """Cell dict {kind, text, crit} for one checklist item on one tower's latest
    inspection. kind ∈ blank|na|np|ok|defect."""
    if inspection is None:
        return {'kind': 'blank', 'text': '', 'crit': ''}
    results = [ir for ir in inspection.item_results.all() if ir.item_id == item.id]
    if not results:
        return {'kind': 'blank', 'text': '', 'crit': ''}

    positional = bool(item.positions)
    order = {p: i for i, p in enumerate(item.positions)} if positional else {}
    crits, defect_segs, statuses = [], [], set()
    for ir in sorted(results, key=lambda r: order.get(r.position, 99)):
        statuses.add(ir.status)
        if ir.status == 'defect':
            entries = list(ir.entries.all())
            crits.extend(e.criticality for e in entries)
            text = ', '.join(
                e.defect.label + (f" ({_answers_text(e.answers, fu_map)})" if e.answers else '')
                for e in entries
            ) or 'Defect'
            defect_segs.append(f"{ir.position}: {text}" if positional and ir.position else text)

    if defect_segs:
        return {'kind': 'defect', 'text': ' | '.join(defect_segs), 'crit': worst_of(crits) if crits else 'minor'}
    if statuses == {'na'}:
        return {'kind': 'na', 'text': 'N/A', 'crit': ''}
    if statuses and statuses <= {'not_provided', 'na'}:
        return {'kind': 'np', 'text': 'Not available', 'crit': ''}
    return {'kind': 'ok', 'text': '✓', 'crit': ''}


def build_register(line, cycle=''):
    """Return {line, cycle, tower_headers, rows} for one line. cycle in
    ('', 'ground_patrol', 'pmi').

    Includes ALL towers on the line (real + VT), ordered by the tower schedule
    (schedule.line_schedule). A VT tower's inspection is taken from the REAL tower
    at the same physical structure (shared structure_key) — so an inspected
    structure reads as inspected on every circuit running through it. VT columns
    are flagged so the UI can badge them."""
    inspection_type = cycle or None
    towers = sched.line_schedule(line, include_virtual=True)

    # Inspection source per tower: real -> itself; VT -> the co-located real tower
    # sharing its structure_key. Resolve all VT structures in one query.
    vt_keys = {t.structure_key for t in towers if t.is_virtual and t.structure_key}
    real_id_by_key = {}
    if vt_keys:
        for rt in (Tower.objects.filter(structure_key__in=vt_keys, is_virtual=False, is_active=True)
                   .only('id', 'structure_key')):
            real_id_by_key.setdefault(rt.structure_key, rt.id)

    def source_id(t):
        return t.id if not t.is_virtual else real_id_by_key.get(t.structure_key)

    latest = tower_status.latest_inspections(
        [sid for t in towers if (sid := source_id(t)) is not None], inspection_type=inspection_type)
    fu_map = {q.key: q for q in FollowUpQuestion.objects.all()}

    tower_headers = []
    for t in towers:
        insp = latest.get(source_id(t))
        tower_headers.append({
            'id': t.id,
            'source_id': source_id(t),
            'tower_number': t.tower_number or str(t.id),
            'tower_type': t.tower_type or '',
            'is_virtual': t.is_virtual,
            'inspected': insp is not None,
            'date': insp.date if insp else None,
            'worst': insp.worst_criticality if insp else 'none',
        })

    def meta_row(sno, label, value_fn):
        return {'group': None, 'sno': sno, 'label': label,
                'cells': [{'kind': 'meta', 'text': value_fn(latest.get(h['source_id']), h), 'crit': ''}
                          for h in tower_headers]}

    rows = [
        meta_row(1, 'Date of inspection', lambda insp, h: h['date'].strftime('%d-%m-%Y') if h['date'] else ''),
        meta_row(2, 'Loc No', lambda insp, h: h['tower_number']),
        meta_row(3, 'Type of tower', lambda insp, h: h['tower_type']),
    ]

    for group in ChecklistItemGroup.objects.prefetch_related('items').order_by('sort_order'):
        items = sorted(group.items.all(), key=lambda i: (i.sort_order, i.sno))
        if not items:
            continue
        rows.append({'group_header': group.label})
        for item in items:
            rows.append({
                'group': group.key, 'sno': item.sno, 'label': item.label,
                'cells': [_item_cell(item, latest.get(h['source_id']), fu_map) for h in tower_headers],
            })

    rows.append(meta_row(None, 'Remarks', lambda insp, h: (insp.remarks if insp else '')))

    return {'line': line, 'cycle': cycle, 'tower_headers': tower_headers, 'rows': rows}


def register_to_xlsx(structure):
    """Serialize a build_register() structure to .xlsx bytes (openpyxl — already
    a project dependency). Defect cells are prefixed with the criticality."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment

    line = structure['line']
    headers = structure['tower_headers']
    wb = Workbook()
    ws = wb.active
    ws.title = 'Register'

    ws.append([f"Line inspection register — {line.name or line.id}"])
    ws['A1'].font = Font(bold=True, size=13)
    ws.append([f"Voltage: {line.voltage}    Cycle: {CYCLE_LABELS.get(structure['cycle'], 'Latest')}"])
    ws.append([])

    ws.append(['S.No', 'Item'] + [f"T-{h['tower_number']}" for h in headers])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)

    for row in structure['rows']:
        if 'group_header' in row:
            ws.append([row['group_header']])
            ws[f'A{ws.max_row}'].font = Font(bold=True, italic=True)
            continue
        values = []
        for c in row['cells']:
            text = c['text']
            if c['kind'] == 'defect' and c['crit']:
                text = f"[{c['crit'].upper()}] {text}"
            values.append(text)
        ws.append([row.get('sno') or '', row['label']] + values)

    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 34
    ws.freeze_panes = 'C5'
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical='top', wrap_text=True)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
