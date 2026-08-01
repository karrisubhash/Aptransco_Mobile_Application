"""Tabular report exports — the Excel / PDF downloads behind the mobile app's
History and Tickets tabs.

One shape of report, two renderings. A report here is deliberately flat: a title,
a few lines of provenance (whose scope, which filters, when it was generated, how
many rows) and then one row per record with fixed columns — the same rows the tab
is showing on screen, in the same order. That is what makes the two renderings
interchangeable, and it is why [build_report] returns a plain structure that
[report_to_xlsx] and [report_to_pdf] each only have to lay out.

    report = build_report(...)          # what to say
    report_to_xlsx(report)              # -> .xlsx bytes
    report_to_pdf(report)               # -> .pdf bytes
    attachment(content, report, 'pdf')  # -> HttpResponse with the download header

Rows come from `inspection_report` / `ticket_report`, which take the *already
scoped* querysets the API views build. Scoping is not repeated here — a report
can only ever contain what its caller's queryset contains, so an export cannot
widen what a viewer is allowed to see.

Read-only; nothing here writes to the database.
"""
import re
from io import BytesIO

from django.http import HttpResponse
from django.utils import timezone

from .models import (
    CRITICALITY_CHOICES, TICKET_SOURCE_CHOICES, TICKET_STATUS_CHOICES,
    WORST_CHOICES, DefectEntry, attach_defect_counts,
)

# How many rows a single export may carry. Higher than the 500 the list
# endpoints page at — an export is a report, and silently handing back the first
# screen of a division's history would read as the whole of it. When a report is
# cut here it says so in its own header (see build_report), so a truncated
# download can never be mistaken for a complete one.
MAX_EXPORT_ROWS = 5000

FORMATS = {
    'xlsx': (
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        'xlsx',
    ),
    'pdf': ('application/pdf', 'pdf'),
}

_WORST_LABEL = dict(WORST_CHOICES)
_CRIT_LABEL = dict(CRITICALITY_CHOICES)
_TICKET_STATUS_LABEL = dict(TICKET_STATUS_CHOICES)
_TICKET_SOURCE_LABEL = dict(TICKET_SOURCE_CHOICES)


# ---------------------------------------------------------------------------
# The report structure
# ---------------------------------------------------------------------------

class Report:
    """A flat tabular report: what to print, independent of how.

    [columns] is a list of `(heading, width)` pairs. The width is in characters
    and is shared by both writers — openpyxl takes it as a column width almost
    directly, and the PDF divides the page across the columns in proportion to
    it — so a column that needs room to be readable gets it in both downloads
    rather than being tuned twice.
    """

    def __init__(self, title, subtitle, columns, rows, notes=(), stem='report'):
        self.title = title
        self.subtitle = subtitle
        self.columns = list(columns)
        self.rows = list(rows)
        self.notes = list(notes)
        self.stem = stem

    @property
    def headings(self):
        return [c[0] for c in self.columns]

    @property
    def widths(self):
        return [c[1] for c in self.columns]

    def filename(self, fmt):
        """`li_history_vijayawada_20260730.xlsx` — safe on every platform the
        app shares to, and sorted by date when several land in one folder."""
        stamp = timezone.localtime().strftime('%Y%m%d_%H%M')
        safe = re.sub(r'[^A-Za-z0-9]+', '_', self.stem).strip('_').lower() or 'report'
        return f'{safe}_{stamp}.{FORMATS[fmt][1]}'


def build_report(title, columns, rows, stem, scope_label='', filters=(), truncated=False):
    """Assembles a [Report], deriving the provenance lines every export carries.

    [filters] is a sequence of `(label, value)` pairs; the empty ones are
    dropped, so a caller can hand over every filter it supports without having
    to work out which ones are set.
    """
    generated = timezone.localtime().strftime('%d %b %Y, %H:%M')
    parts = [p for p in [scope_label, f'Generated {generated}'] if p]
    applied = ' · '.join(f'{label}: {value}' for label, value in filters if value)

    notes = []
    if applied:
        notes.append(applied)
    notes.append(f'{len(rows)} row{"" if len(rows) == 1 else "s"}')
    if truncated:
        # Never a silent cut: a report that hit the ceiling says so on its face,
        # so the reader knows to narrow the filters rather than trusting a total.
        notes.append(
            f'Truncated to the {MAX_EXPORT_ROWS} most recent records — '
            'narrow the filters for a complete report.'
        )
    return Report(
        title=title,
        subtitle=' · '.join(parts),
        columns=columns,
        rows=rows,
        notes=notes,
        stem=stem,
    )


def _date(value, fmt='%d %b %Y'):
    """A date/datetime as a short local string; '' for None."""
    if not value:
        return ''
    if hasattr(value, 'tzinfo') and value.tzinfo is not None:
        value = timezone.localtime(value)
    return value.strftime(fmt)


def _answers(answers):
    """A follow-up answers dict as 'Height: 12 m · Phase: R', matching how the
    app's `answersText` renders the same map on screen."""
    if not isinstance(answers, dict):
        return ''
    out = []
    for key, value in answers.items():
        text = ', '.join(str(v) for v in value) if isinstance(value, list) else str(value)
        if text:
            out.append(f'{key}: {text}')
    return ' · '.join(out)


# ---------------------------------------------------------------------------
# Inspection history
# ---------------------------------------------------------------------------

INSPECTION_COLUMNS = [
    ('Date', 12),
    ('Tower_no', 10),
    ('Line', 30),
    ('Tower Type', 12),
    ('Defect status', 14),
    ('Defects_cnt', 11),
    ('Defects', 40),
    ('Inspector_ID', 14),
    ('Remarks', 30),
]


def defect_labels_by_inspection(inspection_ids):
    """`{inspection_id: 'Bent member (LH top); Paint faded'}` — what was actually
    found, for the report's Defects column.

    One query for the whole page, in the same spirit as [attach_defect_counts]:
    the alternative is prefetching `item_results__entries__defect` on the export
    queryset, which drags every item result — including the thousands of normal
    ones — through Python just to reach the handful that carry a defect.

    Entries come back in a fixed order (item sort order, then position) so the
    same inspection reads the same way in every download of it.
    """
    if not inspection_ids:
        return {}
    rows = (DefectEntry.objects
            .filter(item_result__inspection_id__in=inspection_ids)
            .order_by('item_result__item__sort_order', 'item_result__position', 'id')
            .values_list('item_result__inspection_id', 'defect__label',
                         'item_result__position'))
    out = {}
    for inspection_id, label, position in rows:
        text = f'{label} ({position})' if position else (label or '')
        if text:
            out.setdefault(inspection_id, []).append(text)
    return {k: '; '.join(v) for k, v in out.items()}


def inspection_report(queryset, scope_label='', filters=()):
    """A [Report] over an oversight-scoped Inspection queryset.

    Ordering and scoping are the caller's — this takes the queryset the History
    tab's own list endpoint builds, so the download is the list, not a second
    answer to the same question.

    One row per inspection, with `defect_count` filled in by
    [attach_defect_counts] over just the rows being exported (the same reason
    the list endpoint uses it: annotating the count would group the viewer's
    whole history to label one report), and the defects themselves named
    alongside it by [defect_labels_by_inspection] — a count answers "is this
    tower a problem", the labels answer "what is wrong with it", and a report
    that has to be read away from the app needs both.
    """
    rows = attach_defect_counts(
        queryset.select_related('tower')[:MAX_EXPORT_ROWS + 1]
    )
    truncated = len(rows) > MAX_EXPORT_ROWS
    rows = rows[:MAX_EXPORT_ROWS]
    defects = defect_labels_by_inspection([i.pk for i in rows])

    table = [[
        _date(i.date),
        f'T-{i.tower.tower_number}' if i.tower.tower_number else f'#{i.tower_id}',
        i.tower.line_name or '',
        i.tower.tower_type or '',
        _WORST_LABEL.get(i.worst_criticality, i.worst_criticality),
        i.defect_count,
        defects.get(i.pk, ''),
        i.inspector_employee_id or '',
        i.remarks or '',
    ] for i in rows]

    return build_report(
        title='Inspection History',
        columns=INSPECTION_COLUMNS,
        rows=table,
        stem='li_history',
        scope_label=scope_label,
        filters=filters,
        truncated=truncated,
    )


# ---------------------------------------------------------------------------
# Defect tickets
# ---------------------------------------------------------------------------

TICKET_COLUMNS = [
    ('Raised', 12),
    ('Tower_no', 10),
    ('Line', 26),
    ('Component', 24),
    ('Position', 10),
    ('Defect', 26),
    ('Detail', 22),
    ('Criticality', 12),
    ('Status', 9),
    ('Source', 14),
    ('Raised by', 12),
    ('Closed', 12),
    ('Closed by', 12),
    ('Closure note', 26),
]


def ticket_report(queryset, scope_label='', filters=()):
    """A [Report] over an oversight-scoped DefectTicket queryset.

    Carries the closure columns even when the export is of open tickets only:
    they come out empty, and a fixed column set means the two downloads of the
    same backlog can be compared side by side.
    """
    rows = list(queryset.select_related('tower')[:MAX_EXPORT_ROWS + 1])
    truncated = len(rows) > MAX_EXPORT_ROWS
    rows = rows[:MAX_EXPORT_ROWS]

    table = [[
        _date(t.raised_at),
        f'T-{t.tower.tower_number}' if t.tower.tower_number else f'#{t.tower_id}',
        t.tower.line_name or '',
        t.item_label or '',
        t.position or '',
        t.defect_label or '',
        _answers(t.answers),
        _CRIT_LABEL.get(t.criticality, t.criticality),
        _TICKET_STATUS_LABEL.get(t.status, t.status),
        _TICKET_SOURCE_LABEL.get(t.source, t.source),
        t.raised_by_employee_id or '',
        _date(t.closed_at),
        t.closed_by_employee_id or '',
        t.close_note or '',
    ] for t in rows]

    return build_report(
        title='Defect Tickets',
        columns=TICKET_COLUMNS,
        rows=table,
        stem='li_tickets',
        scope_label=scope_label,
        filters=filters,
        truncated=truncated,
    )


# ---------------------------------------------------------------------------
# Renderings
# ---------------------------------------------------------------------------

def report_to_xlsx(report):
    """Serialize a [Report] to .xlsx bytes (openpyxl — already a dependency for
    register.py's register export).

    The header block sits above the table and the header row is frozen, so
    scrolling a division's history keeps the column names in view.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = report.title[:31]  # Excel's sheet-name ceiling

    ws.append([report.title])
    ws['A1'].font = Font(bold=True, size=14)
    if report.subtitle:
        ws.append([report.subtitle])
        ws[f'A{ws.max_row}'].font = Font(size=10, color='FF5A6675')
    for note in report.notes:
        ws.append([note])
        ws[f'A{ws.max_row}'].font = Font(size=10, color='FF5A6675')
    ws.append([])

    # Read the heading row back off the sheet rather than predicting it: the
    # blank spacer above writes no cells, so `max_row` does not count it while
    # the next append still lands past it — predicting from `max_row + 1` put the
    # frozen pane and the autofilter one row too high.
    ws.append(report.headings)
    head_row = ws.max_row
    header_fill = PatternFill('solid', start_color='FF14243D')  # kBrandPrimary
    for cell in ws[head_row]:
        cell.font = Font(bold=True, color='FFFFFFFF')
        cell.fill = header_fill
        cell.alignment = Alignment(vertical='center', wrap_text=True)

    for row in report.rows:
        ws.append(row)

    for index, width in enumerate(report.widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width
    ws.freeze_panes = f'A{head_row + 1}'
    ws.auto_filter.ref = (
        f'A{head_row}:{get_column_letter(len(report.columns))}'
        f'{head_row + len(report.rows)}'
    )
    for row in ws.iter_rows(min_row=head_row + 1):
        for cell in row:
            cell.alignment = Alignment(vertical='top', wrap_text=True)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


_BODY_FONT_CANDIDATES = [
    # Nirmala UI — Windows' Indic face, and the one that matters here: a field
    # remark or a line name typed in Telugu is entirely ordinary in Andhra
    # Pradesh, and reportlab's built-in Helvetica is Latin-1 only, so those
    # characters come out as empty boxes. It covers Latin too, so it can serve
    # the whole table rather than being switched per cell.
    'C:/Windows/Fonts/Nirmala.ttc',
    # Linux hosts: Noto's Telugu face where it is installed, then DejaVu, which
    # is near-universal and at least covers the punctuation and symbols
    # Latin-1 lacks.
    '/usr/share/fonts/truetype/noto/NotoSansTelugu-Regular.ttf',
    '/usr/share/fonts/noto/NotoSansTelugu-Regular.ttf',
    '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',
]

_BODY_FONT_NAME = 'LiBody'
_body_font_resolved = None


def _body_font():
    """The font name for cells that can hold user-entered text.

    Resolved once per process: the first candidate that registers wins, and if
    none do we stay on Helvetica. Deliberately best-effort — a missing or
    unreadable font file must degrade the typography of a report, never fail the
    download — and only the *data* cells use it. Column headings and the report
    title are app-generated English, so they keep Helvetica-Bold rather than
    forcing a hunt for a matching bold face.
    """
    global _body_font_resolved
    if _body_font_resolved is not None:
        return _body_font_resolved

    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    _body_font_resolved = 'Helvetica'
    for path in _BODY_FONT_CANDIDATES:
        try:
            pdfmetrics.registerFont(TTFont(_BODY_FONT_NAME, path))
        except Exception:
            continue
        _body_font_resolved = _BODY_FONT_NAME
        break
    return _body_font_resolved


def report_to_pdf(report):
    """Render a [Report] as landscape A4 PDF bytes (reportlab).

    Landscape because these tables are wide — a ticket row carries fourteen
    columns, and portrait would either clip them or squeeze them to unreadable.
    The header row repeats on every page (`repeatRows=1`), so page 4 of a
    backlog is still readable on its own.

    Cells that can carry user-entered text are set in [_body_font]; see there for
    why that is not simply Helvetica.
    """
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_LEFT
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
    )

    brand = colors.HexColor('#14243D')   # kBrandPrimary
    ink_soft = colors.HexColor('#5A6675')
    outline = colors.HexColor('#E2E8F1')
    zebra = colors.HexColor('#F4F7FB')   # kSurface

    body = _body_font()
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'LiTitle', parent=styles['Title'],
        fontName='Helvetica-Bold', fontSize=15, leading=18,
        alignment=TA_LEFT, textColor=brand, spaceAfter=2,
    )
    meta_style = ParagraphStyle(
        'LiMeta', parent=styles['Normal'],
        fontName=body, fontSize=8.5, leading=11, textColor=ink_soft,
    )
    head_style = ParagraphStyle(
        'LiHead', parent=styles['Normal'],
        fontName='Helvetica-Bold', fontSize=7.5, leading=9,
        textColor=colors.white,
    )
    cell_style = ParagraphStyle(
        'LiCell', parent=styles['Normal'],
        fontName=body, fontSize=7.5, leading=9.5,
    )

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=10 * mm, rightMargin=10 * mm,
        topMargin=10 * mm, bottomMargin=12 * mm,
        title=report.title,
        author='APTRANSCO Line Inspection',
    )

    story = [Paragraph(report.title, title_style)]
    if report.subtitle:
        story.append(Paragraph(report.subtitle, meta_style))
    for note in report.notes:
        story.append(Paragraph(note, meta_style))
    story.append(Spacer(1, 5 * mm))

    if report.rows:
        # Every cell is a Paragraph so long text (a remark, a closure note) wraps
        # inside its column instead of forcing the table off the page.
        data = [[Paragraph(_escape(h), head_style) for h in report.headings]]
        data += [
            [Paragraph(_escape(value), cell_style) for value in row]
            for row in report.rows
        ]

        # Share the printable width in proportion to the declared column widths,
        # so the same column that is roomy in Excel is roomy here.
        available = doc.width
        total = sum(report.widths) or len(report.widths)
        col_widths = [available * w / total for w in report.widths]

        table = Table(data, colWidths=col_widths, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), brand),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('GRID', (0, 0), (-1, -1), 0.4, outline),
            ('LINEBELOW', (0, 0), (-1, 0), 0.6, brand),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            # Banded rows: on a table this dense, following one row across
            # fourteen columns is the whole readability problem.
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, zebra]),
        ]))
        story.append(table)
    else:
        story.append(Paragraph('No records for these filters.', meta_style))

    doc.build(story, onLaterPages=_page_footer, onFirstPage=_page_footer)
    return buf.getvalue()


def _escape(value):
    """Text safe inside a reportlab Paragraph, whose mini-HTML would otherwise
    read a stray '<' in a remark as markup and fail the whole render."""
    return (str('' if value is None else value)
            .replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;'))


def _page_footer(canvas, doc):
    """'Page n' bottom-right, so a printed report cannot be reassembled wrong."""
    canvas.saveState()
    canvas.setFont('Helvetica', 7.5)
    canvas.setFillColorRGB(0.35, 0.40, 0.46)
    canvas.drawRightString(
        doc.pagesize[0] - 10 * 2.834645,  # 10 mm in points
        7 * 2.834645,
        f'Page {canvas.getPageNumber()}',
    )
    canvas.restoreState()


def render(report, fmt):
    """[report] as bytes in [fmt] ('xlsx' or 'pdf')."""
    if fmt == 'pdf':
        return report_to_pdf(report)
    if fmt == 'xlsx':
        return report_to_xlsx(report)
    raise ValueError(f'Unsupported export format: {fmt!r}')


def attachment(report, fmt):
    """An [HttpResponse] that downloads [report] as [fmt].

    Content-Disposition is what names the file, and the mobile app reads the
    filename straight back out of it — so the name a download lands under on the
    phone is decided here, once, for both formats.
    """
    content_type, _ = FORMATS[fmt]
    resp = HttpResponse(render(report, fmt), content_type=content_type)
    resp['Content-Disposition'] = f'attachment; filename="{report.filename(fmt)}"'
    # The app shows a progress dialog while the file arrives; without a length
    # there is nothing to show progress against.
    resp['Content-Length'] = str(len(resp.content))
    return resp
