"""Phase 3 dashboard tests — oversight scope, status cache, rollups, tier
gating, and the map endpoint's jurisdiction-security invariant.

NOTE ON RUNNING: line_inspection models are routed to the `line_inspection_db`
Postgres connection (the `clear` schema) and the status helper uses Postgres
`DISTINCT ON`, so these tests require a Postgres test database for that
connection — they will not run on the default sqlite alias. `databases` below
tells Django's TestCase to manage transactions on both connections. Until CI
has a Postgres test DB (Phase 5), the equivalent checks are also exercised by
the transactional verification run at the end of Phase 3 development.
"""
from django.test import TestCase, SimpleTestCase, Client, override_settings
from django.utils import timezone
from rest_framework.test import APIRequestFactory

from .. import viewing, jurisdiction, register, schedule
from .. import status as tower_status
from .. import dashboard_views
from ..api.views import MapTowerListView, _create_inspection
from ..management.commands.sync_gis_towers import _parse_circuit_count, _assign_structure_keys
from rest_framework.exceptions import ValidationError
from ..models import (
    Subdivision, Line, Tower, EmployeeCadreSnapshot, RoleAssignment, LineTowerAssignment,
    FieldEECadrePosition, Inspection, ItemResult, DefectEntry, DefectTicket,
    ChecklistItem, Defect, ChecklistItemGroup, FollowUpQuestion, CatalogVersion,
    SuperAdmin, LiloEvent, LiloTowerRemap,
)
from ..services import lilo_matching as lm


class DashboardScopeTests(TestCase):
    databases = {'default', 'line_inspection_db'}

    def setUp(self):
        self.sub = Subdivision.objects.create(name='T SD', circle='TC', division='TD', zone='TZ')
        self.line = Line.objects.create(
            source_layer='t', arcgis_object_id='L1', name='T LINE', voltage='220kV',
            zone='TZ', circle='TC', division='TD', subdivision=self.sub, is_active=True,
            geometry={'type': 'LineString', 'coordinates': [[79.0, 16.0], [79.1, 16.1]]},
        )
        self.t1 = Tower.objects.create(source_layer='t', arcgis_object_id='T1', tower_number='1',
                                       tower_type='Boom', voltage='220kV', line=self.line,
                                       latitude=16.0, longitude=79.0, zone='TZ', circle='TC',
                                       division='TD', is_active=True)
        self.t2 = Tower.objects.create(source_layer='t', arcgis_object_id='T2', tower_number='2',
                                       tower_type='Boom', voltage='220kV', line=self.line,
                                       latitude=16.1, longitude=79.1, zone='TZ', circle='TC',
                                       division='TD', is_active=True)
        EmployeeCadreSnapshot.objects.create(employee_id='SE1', emp_sub_grp='SE', reporting_manager_id='')
        EmployeeCadreSnapshot.objects.create(employee_id='EE1', emp_sub_grp='DE/EE', position_id='P1', reporting_manager_id='SE1')
        EmployeeCadreSnapshot.objects.create(employee_id='AEE1', emp_sub_grp='AE', reporting_manager_id='EE1')
        EmployeeCadreSnapshot.objects.create(employee_id='NONE1', emp_sub_grp='AE', reporting_manager_id='')
        EmployeeCadreSnapshot.objects.create(employee_id='CE1', emp_sub_grp='CE', reporting_manager_id='')
        RoleAssignment.objects.create(employee_id='AEE1', role='FIELD_INSPECTOR', subdivision=self.sub,
                                      assigned_by_employee_id='EE1', is_active=True)

    # --- reporting-hierarchy oversight scope ---
    def test_scope_widens_up_hierarchy(self):
        ids = lambda e: set(viewing.oversight_towers(e).values_list('id', flat=True))
        self.assertEqual(ids('AEE1'), {self.t1.id, self.t2.id})
        self.assertEqual(ids('EE1'), {self.t1.id, self.t2.id})   # via subordinate
        self.assertEqual(ids('SE1'), {self.t1.id, self.t2.id})   # whole subtree
        self.assertEqual(ids('NONE1'), set())

    # --- oversight view must NOT grant edit rights ---
    def test_edit_scope_isolated_from_view_scope(self):
        self.assertEqual(set(jurisdiction.visible_towers('EE1').values_list('id', flat=True)), set())
        self.assertFalse(jurisdiction.can_edit_tower('EE1', self.t1))
        self.assertTrue(jurisdiction.can_edit_tower('AEE1', self.t1))

    # --- ticket close follows the reporting hierarchy; capture does not ---
    def test_oversees_tower_covers_the_subtree(self):
        # The supervisor cannot *capture* on this tower (asserted above) but does
        # oversee it, which is what authorises signing off its defects.
        self.assertTrue(viewing.oversees_tower('EE1', self.t1))
        self.assertTrue(viewing.oversees_tower('SE1', self.t1))
        self.assertTrue(viewing.oversees_tower('AEE1', self.t1))
        self.assertFalse(viewing.oversees_tower('NONE1', self.t1))

    def test_oversees_tower_is_a_superset_of_own_scope(self):
        # Anything an employee may edit, they also oversee — so ORing the two in
        # DefectTicketCloseView can never take away an existing permission.
        for emp in ('AEE1', 'EE1', 'SE1', 'NONE1'):
            for tower in (self.t1, self.t2):
                if jurisdiction.can_edit_tower(emp, tower):
                    self.assertTrue(
                        viewing.oversees_tower(emp, tower),
                        f'{emp} may edit but does not oversee tower {tower.pk}',
                    )

    @override_settings(LINE_INSPECTION_OPEN_INSPECT=True)
    def test_inspection_capture_is_open_to_any_employee(self):
        # Recording an inspection is deliberately open (see
        # settings.LINE_INSPECTION_OPEN_INSPECT): a supervisor with no
        # RoleAssignment of their own may still capture, as may an employee with
        # no assignment at all — who could previously inspect nothing.
        self.assertTrue(jurisdiction.can_inspect_tower('EE1', self.t1))
        self.assertTrue(jurisdiction.can_inspect_tower('NONE1', self.t1))
        for emp in ('EE1', 'NONE1'):
            insp = _create_inspection(emp, {'tower_id': self.t1.id, 'items': []})
            self.assertEqual(insp.inspector_employee_id, emp)

    @override_settings(LINE_INSPECTION_OPEN_INSPECT=False)
    def test_inspection_capture_falls_back_to_own_scope(self):
        # The per-assignment rule is intact, not deleted — turning the flag off
        # restores the original refusal, and ticket close still widens instead.
        from rest_framework.exceptions import PermissionDenied
        self.assertFalse(jurisdiction.can_inspect_tower('EE1', self.t1))
        self.assertTrue(jurisdiction.can_inspect_tower('AEE1', self.t1))
        with self.assertRaises(PermissionDenied):
            _create_inspection('EE1', {'tower_id': self.t1.id, 'items': []})

    @override_settings(LINE_INSPECTION_OPEN_INSPECT=True)
    def test_open_capture_does_not_widen_ticket_close_scope(self):
        # The flag governs capture only. can_edit_tower — which also backs the
        # defect-ticket sign-off — must be unaffected by it.
        self.assertFalse(jurisdiction.can_edit_tower('EE1', self.t1))
        self.assertFalse(jurisdiction.can_edit_tower('NONE1', self.t1))
        self.assertTrue(jurisdiction.can_edit_tower('AEE1', self.t1))

    def test_functional_tiers(self):
        self.assertEqual(viewing.cadre_tier('AEE1'), 'field_user')
        self.assertEqual(viewing.cadre_tier('EE1'), 'supervisor')
        self.assertEqual(viewing.cadre_tier('NONE1'), 'none')
        self.assertTrue(viewing.is_management('CE1'))
        self.assertFalse(viewing.is_management('AEE1'))
        self.assertEqual(viewing.display_label('AEE1'), 'AEE')
        self.assertEqual(viewing.display_label('EE1'), 'EE')

    # --- the management/super-admin all-tower scope reaches lines too ---
    def test_management_scope_covers_lines_and_towers_together(self):
        # CE1 holds no RoleAssignment and has nobody under them, so their
        # reporting subtree is empty — is_management is the only thing granting
        # scope, and it has to grant BOTH halves. The regression this pins: a
        # super admin / CE saw every tower in the KPI and rollup totals while the
        # line list, the map and the register came back empty, because only the
        # tower side of the widening existed.
        self.assertEqual(viewing.visible_employee_ids('CE1'), {'CE1'})
        self.assertEqual(set(viewing.oversight_towers('CE1').values_list('id', flat=True)),
                         {self.t1.id, self.t2.id})
        self.assertEqual(set(viewing.oversight_lines('CE1').values_list('id', flat=True)),
                         {self.line.id})

    def test_super_admin_scope_needs_no_assignment(self):
        # Same invariant by the other route into is_management. 'NONE1' is the
        # employee with no assignment and no subordinates: empty before the
        # SuperAdmin row, everything after it.
        self.assertEqual(set(viewing.oversight_lines('NONE1').values_list('id', flat=True)), set())
        SuperAdmin.objects.create(employee_id='NONE1')
        self.assertEqual(set(viewing.oversight_lines('NONE1').values_list('id', flat=True)),
                         {self.line.id})
        self.assertEqual(set(viewing.oversight_towers('NONE1').values_list('id', flat=True)),
                         {self.t1.id, self.t2.id})

    def test_management_scope_still_excludes_inactive_and_virtual(self):
        # The widening is "all active real towers", not "all rows" — the VT and
        # is_active filters oversight_towers applies must survive it.
        vt = Tower.objects.create(source_layer='t', arcgis_object_id='VT1', tower_number='2A',
                                  tower_type='Boom', voltage='220kV', line=self.line,
                                  latitude=16.05, longitude=79.05, is_active=True, is_virtual=True)
        dead = Tower.objects.create(source_layer='t', arcgis_object_id='T3', tower_number='3',
                                    tower_type='Boom', voltage='220kV', line=self.line,
                                    latitude=16.2, longitude=79.2, is_active=False)
        dead_line = Line.objects.create(source_layer='t', arcgis_object_id='L2', name='DEAD LINE',
                                        voltage='220kV', subdivision=self.sub, is_active=False)
        scope = set(viewing.oversight_towers('CE1').values_list('id', flat=True))
        self.assertNotIn(vt.id, scope)
        self.assertNotIn(dead.id, scope)
        self.assertNotIn(dead_line.id, set(viewing.oversight_lines('CE1').values_list('id', flat=True)))

    def test_management_scope_does_not_grant_capture_rights(self):
        # Oversight is read + sign-off. The own-assignment edit scope is built
        # from jurisdiction.py and must not move with is_management.
        self.assertEqual(set(jurisdiction.visible_towers('CE1').values_list('id', flat=True)), set())
        self.assertFalse(jurisdiction.can_edit_tower('CE1', self.t1))

    # --- status + PMI/Ground-Patrol provision ---
    def test_status_cache_and_inspection_type(self):
        now = timezone.now()
        Inspection.objects.create(tower=self.t1, inspector_employee_id='AEE1', catalog_version=CatalogVersion.current(),
                                  date=now.date(), inspection_type='pmi', worst_criticality='major', saved_at=now)
        tower_status.bump_tower_cache(self.t1, now, 'major', 'pmi')
        self.t1.refresh_from_db()
        self.assertEqual(self.t1.last_worst_criticality, 'major')
        self.assertEqual(self.t1.last_inspection_type, 'pmi')
        self.assertIsNotNone(self.t1.last_inspection_at)
        self.assertEqual(tower_status.latest_inspection_status([self.t1.id, self.t2.id]), {self.t1.id: 'major'})
        self.assertEqual(tower_status.latest_inspection_status([self.t1.id], inspection_type='pmi'), {self.t1.id: 'major'})
        self.assertEqual(tower_status.latest_inspection_status([self.t1.id], inspection_type='ground_patrol'), {})

    # --- management rollups ---
    def test_rollup_counts(self):
        now = timezone.now()
        insp = Inspection.objects.create(tower=self.t1, inspector_employee_id='AEE1', catalog_version=CatalogVersion.current(),
                                         date=now.date(), worst_criticality='major', saved_at=now)
        tower_status.bump_tower_cache(self.t1, now, 'major', None)
        grp = ChecklistItemGroup.objects.create(key='g', label='G')
        item = ChecklistItem.objects.create(group=grp, key='i', sno=1, label='Item')
        defect = Defect.objects.create(item=item, key='d', label='Defect', default_criticality='major')
        for src in ('human_inspection', 'drone_inspection'):
            DefectTicket.objects.create(inspection=insp, tower=self.t1, item=item, item_label='Item', defect=defect,
                                        defect_label='Defect', criticality='major', status='open', source=src,
                                        raised_at=now, raised_by_employee_id='AEE1')
        rows = dashboard_views._rollup('zone', viewing.oversight_towers('SE1'))
        row = next(r for r in rows if r['group'] == 'TZ')
        self.assertEqual((row['total'], row['inspected'], row['pending']), (2, 1, 1))
        self.assertEqual(row['major'], 1)
        self.assertEqual((row['tickets_human'], row['tickets_drone']), (1, 1))

    # --- HTTP tier gating ---
    def _client(self, employee_id):
        c = Client()
        s = c.session
        s['employee_id'] = employee_id
        s['display_name'] = employee_id
        s.save()
        return c

    def test_reports_gating(self):
        FieldEECadrePosition.objects.create(position_id='P1')  # EE1 becomes an admin
        self.assertEqual(self._client('AEE1').get('/inspection/dashboard/reports/').status_code, 403)
        self.assertEqual(self._client('EE1').get('/inspection/dashboard/reports/').status_code, 200)
        self.assertEqual(self._client('AEE1').get('/inspection/dashboard/map/').status_code, 200)

    # --- map endpoint jurisdiction-security invariant ---
    def _map_features(self, employee_id, bbox):
        req = APIRequestFactory().get('/inspection/api/map/towers/?bbox=' + bbox)
        req.session = {'employee_id': employee_id}
        resp = MapTowerListView.as_view()(req)
        resp.render()
        import json
        return json.loads(resp.content)['features']

    def test_map_bbox_and_security(self):
        self.assertEqual(len(self._map_features('AEE1', '70,10,90,20')), 2)
        self.assertEqual(len(self._map_features('AEE1', '80,17,81,18')), 0)   # bbox excludes
        self.assertEqual(len(self._map_features('NONE1', '70,10,90,20')), 0)  # forged bbox, no scope

    # --- line inspection register ---
    def _register_item_cell(self, structure, tower_number, label):
        col = next(i for i, h in enumerate(structure['tower_headers']) if h['tower_number'] == tower_number)
        row = next(r for r in structure['rows'] if r.get('label') == label)
        return row['cells'][col]

    def test_register_matrix_cycle_and_export(self):
        now = timezone.now()
        grp = ChecklistItemGroup.objects.create(key='g', label='Tower parts', sort_order=1)
        item = ChecklistItem.objects.create(group=grp, key='i', sno=5, label='Members bent', sort_order=1)
        q = FollowUpQuestion.objects.create(key='q', question_text='How many', answer_type='number', unit='nos')
        defect = Defect.objects.create(item=item, key='d', label='Bent member', default_criticality='major', ask=['q'])
        # t1: PMI defect + earlier ground-patrol normal; t2: ground-patrol normal
        pmi = Inspection.objects.create(tower=self.t1, inspector_employee_id='AEE1', catalog_version=CatalogVersion.current(),
                                        date=now.date(), inspection_type='pmi', worst_criticality='major', saved_at=now)
        ir = ItemResult.objects.create(inspection=pmi, item=item, status='defect')
        DefectEntry.objects.create(item_result=ir, defect=defect, answers={'q': 3}, criticality='major')
        gp = Inspection.objects.create(tower=self.t1, inspector_employee_id='AEE1', catalog_version=CatalogVersion.current(),
                                       date=now.date(), inspection_type='ground_patrol', worst_criticality='ok', saved_at=now)
        ItemResult.objects.create(inspection=gp, item=item, status='normal')
        t2i = Inspection.objects.create(tower=self.t2, inspector_employee_id='AEE1', catalog_version=CatalogVersion.current(),
                                        date=now.date(), inspection_type='ground_patrol', worst_criticality='ok', saved_at=now)
        ItemResult.objects.create(inspection=t2i, item=item, status='normal')

        st = register.build_register(self.line, '')
        self.assertEqual([h['tower_number'] for h in st['tower_headers']], ['1', '2'])
        c = self._register_item_cell(st, '1', 'Members bent')
        self.assertEqual((c['kind'], c['crit']), ('defect', 'major'))
        self.assertIn('How many: 3', c['text'])
        self.assertEqual(self._register_item_cell(st, '2', 'Members bent')['kind'], 'ok')
        # cycle filter
        self.assertEqual(self._register_item_cell(register.build_register(self.line, 'ground_patrol'), '1', 'Members bent')['kind'], 'ok')
        self.assertEqual(self._register_item_cell(register.build_register(self.line, 'pmi'), '2', 'Members bent')['kind'], 'blank')
        # xlsx export
        from io import BytesIO
        from openpyxl import load_workbook
        ws = load_workbook(BytesIO(register.register_to_xlsx(st))).active
        self.assertEqual([ws.cell(row=4, column=1).value, ws.cell(row=4, column=2).value], ['S.No', 'Item'])

    def test_register_line_scope_enforced(self):
        other_sub = Subdivision.objects.create(name='Other', zone='OZ')
        other_line = Line.objects.create(source_layer='t', arcgis_object_id='L2', name='OTHER',
                                          voltage='132kV', subdivision=other_sub, is_active=True)
        c = self._client('AEE1')
        self.assertEqual(c.get(f'/inspection/dashboard/register/?line={self.line.id}').status_code, 200)
        self.assertEqual(c.get(f'/inspection/dashboard/register/?line={other_line.id}').status_code, 403)


class TowerScheduleProjectionTests(SimpleTestCase):
    """Pure geometry helpers — no DB."""

    def test_line_vertices(self):
        self.assertEqual(schedule.line_vertices({'type': 'LineString', 'coordinates': [[1, 2], [3, 4]]}),
                         [(1, 2), (3, 4)])
        self.assertEqual(schedule.line_vertices(None), [])
        self.assertEqual(len(schedule.line_vertices(
            {'type': 'MultiLineString', 'coordinates': [[[1, 2], [3, 4]], [[5, 6]]]})), 3)

    def test_project_onto_line_orders_along(self):
        verts = [(80.0, 16.0), (80.1, 16.0)]  # west -> east
        a0, _ = schedule.project_onto_line(verts, 80.00, 16.0)
        a1, off1 = schedule.project_onto_line(verts, 80.05, 16.0)
        a2, _ = schedule.project_onto_line(verts, 80.10, 16.0)
        self.assertTrue(a0 < a1 < a2)          # along-distance increases eastward
        self.assertAlmostEqual(off1, 0.0, places=1)  # on the line -> ~zero offset
        self.assertEqual(schedule.project_onto_line([(1, 1)], 1, 1), (None, None))  # degenerate


class RegisterVTCompletenessTests(TestCase):
    databases = {'default', 'line_inspection_db'}

    def test_register_includes_vt_with_colocated_status(self):
        sub = Subdivision.objects.create(name='C', zone='Z')
        line_a = Line.objects.create(source_layer='v', arcgis_object_id='A', name='DC Ckt-1',
                                     voltage='132kV', subdivision=sub, is_active=True)
        line_b = Line.objects.create(source_layer='v', arcgis_object_id='B', name='DC Ckt-2',
                                     voltage='132kV', subdivision=sub, is_active=True)
        key = '132kV:16.10000:80.10000'
        real = Tower.objects.create(source_layer='v', arcgis_object_id='R', tower_number='5', voltage='132kV',
                                    line=line_a, is_virtual=False, structure_key=key, line_sequence=1, is_active=True)
        vt = Tower.objects.create(source_layer='v', arcgis_object_id='VT', tower_number='VT 5', voltage='132kV',
                                  line=line_b, is_virtual=True, structure_key=key, line_sequence=1, is_active=True)
        # inspect the real (structure) tower
        Inspection.objects.create(tower=real, inspector_employee_id='x', catalog_version=CatalogVersion.current(),
                                  date='2026-06-01', worst_criticality='major', saved_at='2026-06-01T00:00:00Z')
        tower_status.bump_tower_cache(real, timezone.now(), 'major', None)

        st = register.build_register(line_b, '')  # register for the VT-carrying circuit
        headers = st['tower_headers']
        self.assertEqual([h['tower_number'] for h in headers], ['VT 5'])   # VT column is present
        self.assertTrue(headers[0]['is_virtual'])
        self.assertTrue(headers[0]['inspected'])                            # status borrowed from co-located real
        self.assertEqual(headers[0]['worst'], 'major')


class FromToRangeTests(TestCase):
    databases = {'default', 'line_inspection_db'}

    def setUp(self):
        self.sub = Subdivision.objects.create(name='RC', zone='RZ')
        self.line = Line.objects.create(source_layer='r', arcgis_object_id='RL', name='Range Line',
                                        voltage='132kV', subdivision=self.sub, is_active=True)
        self.t = []
        for seq in range(1, 6):  # towers seq 1..5
            self.t.append(Tower.objects.create(
                source_layer='r', arcgis_object_id=f'T{seq}', tower_number=str(seq), voltage='132kV',
                line=self.line, is_virtual=False, line_sequence=seq, latitude=16.0 + seq / 100, longitude=80.0, is_active=True))
        for emp in ('SE', 'AEEA', 'AEEB'):
            EmployeeCadreSnapshot.objects.create(employee_id=emp, emp_sub_grp='AE', reporting_manager_id='SE' if emp != 'SE' else '')

    def test_range_restricts_jurisdiction_and_splits(self):
        # A gets towers 1-3, B gets 4-5 on the same line
        LineTowerAssignment.objects.create(employee_id='AEEA', line=self.line, from_tower=self.t[0], to_tower=self.t[2],
                                           assigned_by_employee_id='SE', is_active=True)
        LineTowerAssignment.objects.create(employee_id='AEEB', line=self.line, from_tower=self.t[3], to_tower=self.t[4],
                                           assigned_by_employee_id='SE', is_active=True)
        a_ids = set(jurisdiction.visible_towers('AEEA').values_list('id', flat=True))
        b_ids = set(jurisdiction.visible_towers('AEEB').values_list('id', flat=True))
        self.assertEqual(a_ids, {self.t[0].id, self.t[1].id, self.t[2].id})
        self.assertEqual(b_ids, {self.t[3].id, self.t[4].id})
        self.assertTrue(jurisdiction.can_edit_tower('AEEA', self.t[1]))
        self.assertFalse(jurisdiction.can_edit_tower('AEEA', self.t[4]))   # outside A's range

    def test_supervisor_oversight_unions_ranges(self):
        LineTowerAssignment.objects.create(employee_id='AEEA', line=self.line, from_tower=self.t[0], to_tower=self.t[2],
                                           assigned_by_employee_id='SE', is_active=True)
        LineTowerAssignment.objects.create(employee_id='AEEB', line=self.line, from_tower=self.t[3], to_tower=self.t[4],
                                           assigned_by_employee_id='SE', is_active=True)
        se_ids = set(viewing.oversight_towers('SE').values_list('id', flat=True))
        self.assertEqual(se_ids, {t.id for t in self.t})   # SE sees both subordinates' stretches


class TowerInspectionHistoryTests(TestCase):
    databases = {'default', 'line_inspection_db'}

    def setUp(self):
        from django.db.models import Count
        self.Count = Count
        self.sub = Subdivision.objects.create(name='H', zone='HZ')
        self.line = Line.objects.create(source_layer='h', arcgis_object_id='HL', name='Hist Line',
                                        voltage='132kV', subdivision=self.sub, is_active=True)
        self.tw = Tower.objects.create(source_layer='h', arcgis_object_id='HT', tower_number='7', voltage='132kV',
                                       line=self.line, is_virtual=False, line_sequence=1, latitude=16.0, longitude=80.0, is_active=True)
        self.other = Tower.objects.create(source_layer='h', arcgis_object_id='HO', tower_number='9', voltage='132kV',
                                          is_virtual=False, is_active=True)  # no line -> out of scope
        grp = ChecklistItemGroup.objects.create(key='g', label='G', sort_order=1)
        item = ChecklistItem.objects.create(group=grp, key='i', sno=5, label='Members bent', sort_order=1)
        defect = Defect.objects.create(item=item, key='d', label='Bent member', default_criticality='major')
        now = timezone.now()
        self.gp = Inspection.objects.create(tower=self.tw, inspector_employee_id='HE', catalog_version=CatalogVersion.current(),
                                            date='2026-01-10', inspection_type='ground_patrol', worst_criticality='ok',
                                            saved_at=now.replace(microsecond=0))
        ItemResult.objects.create(inspection=self.gp, item=item, status='normal')
        self.pmi = Inspection.objects.create(tower=self.tw, inspector_employee_id='HE', catalog_version=CatalogVersion.current(),
                                             date='2026-05-20', inspection_type='pmi', worst_criticality='major', saved_at=now)
        ir = ItemResult.objects.create(inspection=self.pmi, item=item, status='defect')
        DefectEntry.objects.create(item_result=ir, defect=defect, answers={'count': 2}, criticality='major', note='left leg')
        EmployeeCadreSnapshot.objects.create(employee_id='HE', emp_sub_grp='AE', reporting_manager_id='')
        RoleAssignment.objects.create(employee_id='HE', role='FIELD_INSPECTOR', subdivision=self.sub,
                                      assigned_by_employee_id='x', is_active=True)

    def test_history_queryset_ordering_and_defect_count(self):
        rows = list(Inspection.objects.filter(tower=self.tw)
                    .annotate(defect_count=self.Count('item_results__entries')).order_by('-saved_at'))
        self.assertEqual([r.id for r in rows], [self.pmi.id, self.gp.id])   # newest first
        self.assertEqual([r.defect_count for r in rows], [1, 0])

    def _client(self, eid):
        c = Client(); s = c.session; s['employee_id'] = eid; s['display_name'] = eid; s.save()
        return c

    def test_per_tower_history_renders_with_drilldown(self):
        body = self._client('HE').get(f'/inspection/dashboard/inspections/?tower={self.tw.id}').content.decode()
        for token in ('Members bent', 'Bent member', 'left leg', '20 May 2026', '10 Jan 2026'):
            self.assertIn(token, body)
        self.assertEqual(body.count('<details'), 2)
        self.assertLess(body.index('20 May 2026'), body.index('10 Jan 2026'))   # newest first

    def test_cycle_filter_and_out_of_scope(self):
        pmi_only = self._client('HE').get(f'/inspection/dashboard/inspections/?tower={self.tw.id}&cycle=pmi').content.decode()
        self.assertEqual(pmi_only.count('<details'), 1)
        self.assertNotIn('10 Jan 2026', pmi_only)
        oos = self._client('HE').get(f'/inspection/dashboard/inspections/?tower={self.other.id}').content.decode()
        self.assertIn("isn't within your jurisdiction", oos)


class LiloReconcileTests(TestCase):
    databases = {'default', 'line_inspection_db'}
    GEOM = {'type': 'LineString', 'coordinates': [[80.0, 16.0], [80.0, 16.03]]}

    def setUp(self):
        self.sub = Subdivision.objects.create(name='LSD', zone='Z')
        self.L = Line.objects.create(source_layer='l', arcgis_object_id='OLD', name='OLD', voltage='132kV', subdivision=self.sub, is_active=False)
        self.L1 = Line.objects.create(source_layer='l', arcgis_object_id='NEW', name='NEW A-C', voltage='132kV', subdivision=self.sub, is_active=True, geometry=self.GEOM)
        self.O = {i: Tower.objects.create(source_layer='l', arcgis_object_id=f'O{i}', tower_number=str(i), voltage='132kV',
                                          line=self.L, line_name='OLD', is_active=False, latitude=16.0 + i * 0.005, longitude=80.0)
                  for i in (1, 2, 3)}
        self.insp = Inspection.objects.create(tower=self.O[3], inspector_employee_id='X', catalog_version=CatalogVersion.current(),
                                              date='2026-05-01', worst_criticality='major', saved_at=timezone.now())
        self.N = {i: Tower.objects.create(source_layer='l', arcgis_object_id=f'N{i}', tower_number=str(i), voltage='132kV',
                                          line=self.L1, subdivision=self.sub, is_active=True, latitude=16.0 + i * 0.005, longitude=80.0)
                  for i in (1, 2, 3)}
        self.loop = Tower.objects.create(source_layer='l', arcgis_object_id='NL', tower_number='LOOP', voltage='132kV',
                                         line=self.L1, subdivision=self.sub, is_active=True, latitude=16.02, longitude=80.05)
        self.event = LiloEvent.objects.create(old_line=self.L, performed_by_employee_id='SA')
        self.event.new_lines.add(self.L1)

    def test_propose_classifies(self):
        by = {p['new_tower'].tower_number: p for p in lm.propose_matches(self.event)['pairs']}
        self.assertEqual(by['1']['method'], 'auto')
        self.assertEqual(by['1']['old_tower'].id, self.O[1].id)
        self.assertEqual(by['LOOP']['method'], 'new')

    def test_apply_preserves_history(self):
        decisions = {self.N[1].id: self.O[1].id, self.N[2].id: self.O[2].id, self.N[3].id: self.O[3].id, self.loop.id: None}
        summary = lm.apply_reconcile(self.event, decisions)
        self.assertEqual(summary, {'matched': 3, 'new': 1})
        self.insp.refresh_from_db(); self.O[3].refresh_from_db()
        self.assertEqual(self.insp.tower_id, self.O[3].id)                    # history preserved
        self.assertTrue(self.O[3].is_active and self.O[3].line_id == self.L1.id)
        self.assertEqual(self.O[3].arcgis_object_id, 'N3')                    # adopted new object_id
        self.assertFalse(Tower.objects.filter(id__in=[self.N[1].id, self.N[2].id, self.N[3].id]).exists())  # duplicates gone
        self.assertTrue(Tower.objects.filter(id=self.loop.id, is_active=True).exists())
        self.assertEqual(sorted(LiloTowerRemap.objects.filter(event=self.event).values_list('method', flat=True)),
                         ['auto', 'auto', 'auto', 'new'])
        self.O[3].refresh_from_db()
        self.assertIsNotNone(self.O[3].line_sequence)                         # schedule rebuilt

    def test_detect_churn(self):
        Lc = Line.objects.create(source_layer='l', arcgis_object_id='CL', name='CHURN', voltage='220kV', subdivision=self.sub, is_active=True)
        strand = Tower.objects.create(source_layer='l', arcgis_object_id='SD', tower_number='9', voltage='220kV', line=Lc, is_active=False, latitude=17.0, longitude=81.0)
        Inspection.objects.create(tower=strand, inspector_employee_id='X', catalog_version=CatalogVersion.current(), date='2026-01-01', worst_criticality='ok', saved_at=timezone.now())
        twin = Tower.objects.create(source_layer='l', arcgis_object_id='TW', tower_number='9A', voltage='220kV', line=Lc, is_active=True, latitude=17.0, longitude=81.0)
        self.assertTrue(any(f['stranded'].id == strand.id and f['twin'].id == twin.id for f in lm.detect_churn()))

    def test_lilo_screen_is_super_admin_only(self):
        SuperAdmin.objects.create(employee_id='SA')
        EmployeeCadreSnapshot.objects.create(employee_id='NA', emp_sub_grp='AE', reporting_manager_id='')
        def cli(eid):
            c = Client(); s = c.session; s['employee_id'] = eid; s['display_name'] = eid; s.save(); return c
        self.assertEqual(cli('SA').get('/inspection/admin/lilo/').status_code, 200)
        self.assertEqual(cli('NA').get('/inspection/admin/lilo/').status_code, 403)


class CircuitParseTests(TestCase):
    """Pure-function tests for the sync helpers — no DB needed."""

    def test_parse_circuit_count(self):
        self.assertEqual(_parse_circuit_count('DCT, 2C, ACFR Panther', 'DC'), 2)
        self.assertEqual(_parse_circuit_count('SCT, 1C, Single ACSR Lynx', 'SC'), 1)
        self.assertEqual(_parse_circuit_count('MCT, 4C, Panther', 'MC'), 4)
        self.assertEqual(_parse_circuit_count('Boom', 'DC'), None)          # terminal structure
        self.assertEqual(_parse_circuit_count('', 'MC/DC'), 4)              # circuit_type fallback
        self.assertEqual(_parse_circuit_count('', ''), None)

    def test_structure_linking_by_colocation(self):
        class T:
            def __init__(self): self.structure_key = None
        real, vt, far = T(), T(), T()
        entries = [
            {'tower': real, 'lat': 16.55571, 'lng': 81.54668, 'is_virtual': False},
            {'tower': vt, 'lat': 16.55571, 'lng': 81.54668, 'is_virtual': True},   # co-located
            {'tower': far, 'lat': 16.90000, 'lng': 81.90000, 'is_virtual': True},  # orphan VT
        ]
        _assign_structure_keys(entries, '132kV')
        self.assertTrue(real.structure_key)
        self.assertEqual(vt.structure_key, real.structure_key)   # VT snaps to its real tower
        self.assertNotEqual(far.structure_key, real.structure_key)


class TowerStructureTests(TestCase):
    databases = {'default', 'line_inspection_db'}

    def setUp(self):
        self.sub = Subdivision.objects.create(name='VT SD', zone='VZ')
        self.l1 = Line.objects.create(source_layer='v', arcgis_object_id='L1', name='DC Ckt-1',
                                      voltage='132kV', circuit_type='DC', subdivision=self.sub, is_active=True)
        self.l2 = Line.objects.create(source_layer='v', arcgis_object_id='L2', name='DC Ckt-2',
                                      voltage='132kV', circuit_type='DC', subdivision=self.sub, is_active=True)
        key = '132kV:16.10000:80.10000'
        self.real = Tower.objects.create(source_layer='v', arcgis_object_id='R', tower_number='42', voltage='132kV',
                                         line=self.l1, is_virtual=False, structure_key=key, latitude=16.1, longitude=80.1, is_active=True)
        self.vt = Tower.objects.create(source_layer='v', arcgis_object_id='V', tower_number='VT 42', voltage='132kV',
                                       line=self.l2, is_virtual=True, structure_key=key, latitude=16.1, longitude=80.1, is_active=True)
        EmployeeCadreSnapshot.objects.create(employee_id='VE', emp_sub_grp='AE', reporting_manager_id='')
        RoleAssignment.objects.create(employee_id='VE', role='FIELD_INSPECTOR', subdivision=self.sub,
                                      assigned_by_employee_id='x', is_active=True)

    def test_scopes_exclude_vt(self):
        self.assertEqual(set(viewing.oversight_towers('VE').values_list('id', flat=True)), {self.real.id})
        self.assertEqual(set(jurisdiction.visible_towers('VE').values_list('id', flat=True)), {self.real.id})

    def test_circuits_at_returns_all_colocated_lines(self):
        self.assertEqual({l.id for l in viewing.circuits_at(self.real)}, {self.l1.id, self.l2.id})

    def test_register_lists_real_only(self):
        st = register.build_register(self.l1, '')
        self.assertEqual([h['tower_number'] for h in st['tower_headers']], ['42'])

    def test_inspecting_vt_rejected(self):
        from datetime import date
        with self.assertRaises(ValidationError):
            _create_inspection('VE', {'tower': self.vt.id, 'date': date.today(), 'item_results': []})
