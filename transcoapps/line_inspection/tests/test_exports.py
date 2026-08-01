"""Report-download tests — the Excel / PDF exports behind the app's History and
Tickets tabs (line_inspection/exports.py + the two API export endpoints).

Two classes, split by what they need:

  * [ReportStructureTests] is a SimpleTestCase — the report structure, the
    provenance lines and both writers are pure, so they are testable with no
    database at all. These run anywhere.
  * [ExportEndpointTests] hits the endpoints and so needs the Postgres test
    database for the `line_inspection_db` connection (the `clear` schema), same
    as test_dashboard.py — see the note at the top of that file.

The invariant worth naming: an export must contain exactly what the tab's own
list endpoint returns, no more. `test_export_matches_the_list_endpoint` and
`test_export_cannot_escape_the_viewers_scope` are the two halves of that.
"""
from io import BytesIO

from django.test import Client, SimpleTestCase, TestCase
from django.utils import timezone

from .. import exports
from ..models import (
    CatalogVersion, ChecklistItem, ChecklistItemGroup, Defect, DefectEntry,
    DefectTicket, EmployeeCadreSnapshot, Inspection, ItemResult, Line,
    RoleAssignment, Subdivision, Tower,
)


def _sheet_and_head(content):
    """The worksheet out of .xlsx bytes, plus the 1-indexed row its table
    headings are on.

    Found by scanning rather than assumed: the provenance block above the table
    varies in height with how many filters a report was taken under.
    """
    from openpyxl import load_workbook

    ws = load_workbook(BytesIO(content)).active
    for r in range(1, ws.max_row + 1):
        first = ws.cell(row=r, column=1).value
        if first in ('Date', 'Raised'):
            return ws, r
    raise AssertionError('no heading row in the exported sheet')


def _rows_by_heading(ws, head):
    """The table body as `[{heading: value}]`.

    Keyed by heading, never by position: the first version of these tests read
    `row[5]` and `row[8]`, and renaming one column plus inserting another
    silently moved what every one of those assertions was checking.
    """
    headings = [c.value for c in ws[head]]
    return [
        dict(zip(headings, [c.value for c in ws[r]]))
        for r in range(head + 1, ws.max_row + 1)
    ]


class ReportStructureTests(SimpleTestCase):
    """The report structure and both writers — no database."""

    def _report(self, rows=(), **kw):
        return exports.build_report(
            title='Inspection History',
            columns=exports.INSPECTION_COLUMNS,
            rows=list(rows),
            stem='li_history',
            **kw,
        )

    def test_provenance_lines_name_the_scope_filters_and_count(self):
        report = self._report(
            rows=[['24 Jul 2026', 'T-1', 'L', 'DA', '', 'Major', 1, 'AEE1', '']],
            scope_label='AEE1 · AEE · your oversight scope',
            filters=[('Line', 'T LINE'), ('Tower', ''), ('Inspector', '')],
        )
        self.assertIn('AEE1 · AEE · your oversight scope', report.subtitle)
        self.assertIn('Generated', report.subtitle)
        # Set filters are named; unset ones are dropped rather than printed empty.
        self.assertIn('Line: T LINE', report.notes)
        self.assertNotIn('Tower: ', ' '.join(report.notes))
        self.assertIn('1 row', report.notes)

    def test_row_count_is_plural_beyond_one(self):
        self.assertIn('2 rows', self._report(rows=[[], []]).notes)
        self.assertIn('0 rows', self._report().notes)

    def test_truncation_is_never_silent(self):
        """A report cut at the ceiling has to say so on its face — otherwise the
        first 5000 rows of a division's history read as the whole of it."""
        plain = self._report(rows=[[]])
        cut = self._report(rows=[[]], truncated=True)
        self.assertFalse([n for n in plain.notes if 'runcated' in n])
        self.assertTrue([n for n in cut.notes if 'runcated' in n])
        self.assertIn(str(exports.MAX_EXPORT_ROWS), ' '.join(cut.notes))

    def test_filename_carries_the_stem_and_extension(self):
        report = self._report()
        self.assertTrue(report.filename('xlsx').startswith('li_history_'))
        self.assertTrue(report.filename('xlsx').endswith('.xlsx'))
        self.assertTrue(report.filename('pdf').endswith('.pdf'))

    def test_answers_render_like_the_app_does(self):
        self.assertEqual(exports._answers({'How many': 3}), 'How many: 3')
        self.assertEqual(
            exports._answers({'Which': ['R', 'Y'], 'Height': '12 m'}),
            'Which: R, Y · Height: 12 m',
        )
        self.assertEqual(exports._answers(None), '')
        self.assertEqual(exports._answers({}), '')

    def test_escape_keeps_angle_brackets_out_of_the_pdf_markup(self):
        """reportlab reads a Paragraph as mini-HTML, so an unescaped '<' in a
        remark would fail the whole render rather than print."""
        self.assertEqual(exports._escape('a < b & c > d'), 'a &lt; b &amp; c &gt; d')
        self.assertEqual(exports._escape(None), '')

    def test_history_headings_are_the_agreed_wording(self):
        """The column names are what the reader of the report sees, so they are
        pinned rather than left to whatever the code happens to say. No 'Cycle':
        every tower is inspected under both cycles, and the column said nothing
        the Date did not."""
        self.assertEqual(
            [c[0] for c in exports.INSPECTION_COLUMNS],
            ['Date', 'Tower_no', 'Line', 'Tower Type', 'Defect status',
             'Defects_cnt', 'Defects', 'Inspector_ID', 'Remarks'],
        )
        self.assertEqual([c[0] for c in exports.TICKET_COLUMNS][:2],
                         ['Raised', 'Tower_no'])

    # --- the two writers ---

    def _rows(self):
        return [
            ['24 Jul 2026', 'T-1', 'T LINE 220kV', 'Boom', 'Major', 3,
             'Bent member (LH top); Paint faded', 'AEE1',
             'Disc cracked < LH top > & nest'],
            ['23 Jul 2026', 'T-2', 'T LINE 220kV', 'Boom', 'Normal', 0, '',
             'AEE1', ''],
        ]

    def test_xlsx_has_a_header_block_a_frozen_table_and_the_rows(self):
        report = self._report(rows=self._rows(), scope_label='AEE1')
        ws, head = _sheet_and_head(exports.report_to_xlsx(report))
        self.assertEqual(ws['A1'].value, 'Inspection History')
        self.assertEqual(
            [c.value for c in ws[head]], [c[0] for c in exports.INSPECTION_COLUMNS],
        )
        self.assertEqual(ws.freeze_panes, f'A{head + 1}')

        body = _rows_by_heading(ws, head)
        self.assertEqual([r['Tower_no'] for r in body], ['T-1', 'T-2'])
        self.assertEqual(body[0]['Defects_cnt'], 3)
        self.assertEqual(body[0]['Defects'], 'Bent member (LH top); Paint faded')
        # Verbatim text, brackets and all — the spreadsheet is the copy to grep.
        self.assertEqual(body[0]['Remarks'], 'Disc cracked < LH top > & nest')

    def test_pdf_renders_and_survives_an_empty_table(self):
        full = exports.report_to_pdf(self._report(rows=self._rows()))
        self.assertTrue(full.startswith(b'%PDF-'))
        empty = exports.report_to_pdf(self._report())
        self.assertTrue(empty.startswith(b'%PDF-'))

    def test_pdf_renders_non_latin_text(self):
        """A Telugu remark must not sink the export. Helvetica is Latin-1 only,
        so [exports._body_font] switches the data cells onto a Unicode face
        where the host has one — and falls back rather than failing where it
        does not."""
        rows = [['24 Jul 2026', 'T-1', 'ఉప్పలపాడు', 'Boom', '', 'Major', 1, 'AEE1',
                 'పగిలిన డిస్క్ — 50%']]
        self.assertTrue(exports.report_to_pdf(self._report(rows=rows)).startswith(b'%PDF-'))

    def test_ticket_columns_are_fixed_across_status_filters(self):
        """Open-only and closed-only exports carry the same columns, so two
        downloads of one backlog can be read side by side."""
        self.assertIn('Closure note', [c[0] for c in exports.TICKET_COLUMNS])
        self.assertIn('Criticality', [c[0] for c in exports.TICKET_COLUMNS])

    def test_render_rejects_an_unknown_format(self):
        with self.assertRaises(ValueError):
            exports.render(self._report(), 'docx')


class ExportEndpointTests(TestCase):
    """The two download endpoints, end to end. Needs the Postgres test DB for
    `line_inspection_db` (see the module docstring)."""

    databases = {'default', 'line_inspection_db'}

    XLSX = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    HISTORY = '/inspection/api/line-inspections/export/'
    TICKETS = '/inspection/api/tickets/export/'

    def setUp(self):
        self.sub = Subdivision.objects.create(name='X SD', circle='XC', division='XD', zone='XZ')
        self.line = Line.objects.create(
            source_layer='t', arcgis_object_id='XL1', name='X LINE', voltage='220kV',
            zone='XZ', circle='XC', division='XD', subdivision=self.sub, is_active=True,
        )
        self.tower = Tower.objects.create(
            source_layer='t', arcgis_object_id='XT1', tower_number='7', tower_type='Boom',
            voltage='220kV', line=self.line, line_name='X LINE', latitude=16.0, longitude=79.0,
            zone='XZ', circle='XC', division='XD', is_active=True,
        )
        EmployeeCadreSnapshot.objects.create(employee_id='AEE1', emp_sub_grp='AE', reporting_manager_id='')
        EmployeeCadreSnapshot.objects.create(employee_id='NONE1', emp_sub_grp='AE', reporting_manager_id='')
        RoleAssignment.objects.create(
            employee_id='AEE1', role='FIELD_INSPECTOR', subdivision=self.sub,
            assigned_by_employee_id='AEE1', is_active=True,
        )

        now = timezone.now()
        group = ChecklistItemGroup.objects.create(key='g', label='Tower parts', sort_order=1)
        item = ChecklistItem.objects.create(group=group, key='i', sno=5, label='Members bent', sort_order=1)
        defect = Defect.objects.create(item=item, key='d', label='Bent member', default_criticality='major')

        self.inspection = Inspection.objects.create(
            tower=self.tower, inspector_employee_id='AEE1',
            catalog_version=CatalogVersion.current(), date=now.date(),
            inspection_type='pmi', worst_criticality='major', saved_at=now,
            remarks='Bent leg on the A face',
        )
        result = ItemResult.objects.create(inspection=self.inspection, item=item, status='defect')
        DefectEntry.objects.create(item_result=result, defect=defect, answers={'How many': 3}, criticality='major')

        self.open_ticket = DefectTicket.objects.create(
            inspection=self.inspection, tower=self.tower, item=item, item_label='Members bent',
            defect=defect, defect_label='Bent member', answers={'How many': 3},
            criticality='major', status='open', raised_at=now, raised_by_employee_id='AEE1',
        )
        self.closed_ticket = DefectTicket.objects.create(
            inspection=self.inspection, tower=self.tower, item=item, item_label='Members bent',
            defect=defect, defect_label='Bent member', criticality='minor', status='closed',
            raised_at=now, raised_by_employee_id='AEE1', closed_at=now,
            closed_by_employee_id='AEE1', close_note='Attended and rectified',
        )

    def _client(self, employee_id):
        c = Client()
        s = c.session
        s['employee_id'] = employee_id
        s['display_name'] = employee_id
        s.save()
        return c

    def _table(self, response):
        """The exported table body as `[{heading: value}]`."""
        ws, head = _sheet_and_head(response.content)
        return _rows_by_heading(ws, head)

    def test_an_explicit_format_is_not_swallowed_by_drf(self):
        """`?format=` is DRF's own renderer-selection parameter, and its default
        negotiation 404s when no renderer matches the value — so every download
        the app asked for returned 404 while the endpoint looked fine without the
        parameter. [FileDownloadNegotiation] is what makes these four pass; the
        no-parameter default is deliberately checked alongside them, because it
        was the one case that worked before."""
        client = self._client('AEE1')
        for url in (self.HISTORY, self.TICKETS):
            for query in ('', '?format=xlsx', '?format=pdf'):
                with self.subTest(url=url, query=query):
                    self.assertEqual(client.get(url + query).status_code, 200)

    def test_history_xlsx_downloads_as_a_named_attachment(self):
        resp = self._client('AEE1').get(self.HISTORY + '?format=xlsx')
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp['Content-Type'], self.XLSX)
        self.assertIn('attachment; filename="li_history_', resp['Content-Disposition'])
        self.assertTrue(resp['Content-Disposition'].rstrip('"').endswith('.xlsx'))
        # Content-Length is what the app's progress dialog measures against.
        self.assertEqual(int(resp['Content-Length']), len(resp.content))
        self.assertTrue(resp.content.startswith(b'PK'))  # a real zip container

    def test_history_pdf_downloads_as_a_pdf(self):
        resp = self._client('AEE1').get(self.HISTORY + '?format=pdf')
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp['Content-Type'], 'application/pdf')
        self.assertIn('.pdf"', resp['Content-Disposition'])
        self.assertTrue(resp.content.startswith(b'%PDF-'))

    def test_history_rows_carry_the_inspection(self):
        rows = self._table(self._client('AEE1').get(self.HISTORY + '?format=xlsx'))
        self.assertEqual(len(rows), 1)
        row = rows[0]
        self.assertEqual(row['Tower_no'], 'T-7')
        self.assertEqual(row['Line'], 'X LINE')
        self.assertEqual(row['Tower Type'], 'Boom')
        self.assertEqual(row['Defect status'], 'Major')
        self.assertEqual(row['Defects_cnt'], 1)  # from attach_defect_counts
        self.assertEqual(row['Inspector_ID'], 'AEE1')
        self.assertEqual(row['Remarks'], 'Bent leg on the A face')

    def test_history_names_the_defects_beside_the_count(self):
        """The count says how bad a tower is; the Defects column says what is
        wrong with it. A report read away from the app needs the second."""
        rows = self._table(self._client('AEE1').get(self.HISTORY + '?format=xlsx'))
        self.assertEqual(rows[0]['Defects'], 'Bent member')

    def test_defect_labels_carry_the_position_and_join_in_order(self):
        """A positional item is only identifiable with its position, and two
        defects on one tower have to arrive as one readable cell."""
        item = ChecklistItem.objects.get(key='i')
        second = Defect.objects.create(
            item=item, key='d2', label='Paint faded', default_criticality='minor',
        )
        result = ItemResult.objects.create(
            inspection=self.inspection, item=item, position='LH top', status='defect',
        )
        DefectEntry.objects.create(
            item_result=result, defect=second, criticality='minor',
        )
        rows = self._table(self._client('AEE1').get(self.HISTORY + '?format=xlsx'))
        self.assertEqual(rows[0]['Defects_cnt'], 2)
        self.assertEqual(rows[0]['Defects'], 'Bent member; Paint faded (LH top)')

    def test_tickets_export_honours_the_status_filter(self):
        client = self._client('AEE1')
        self.assertEqual(len(self._table(client.get(self.TICKETS + '?format=xlsx'))), 2)
        open_rows = self._table(client.get(self.TICKETS + '?status=open&format=xlsx'))
        self.assertEqual(len(open_rows), 1)
        self.assertEqual(open_rows[0]['Status'], 'Open')
        self.assertEqual(open_rows[0]['Tower_no'], 'T-7')
        closed_rows = self._table(client.get(self.TICKETS + '?status=closed&format=xlsx'))
        self.assertEqual(len(closed_rows), 1)
        self.assertEqual(closed_rows[0]['Closure note'], 'Attended and rectified')

    def test_export_matches_the_list_endpoint(self):
        """The download and the screen must be the same answer. Both go through
        the same scope helper, and this is what holds them there."""
        client = self._client('AEE1')
        listed = client.get('/inspection/api/tickets/?status=open').json()
        exported = self._table(client.get(self.TICKETS + '?status=open&format=xlsx'))
        self.assertEqual(len(exported), len(listed))
        self.assertEqual(
            [r['Tower_no'] for r in exported],
            [f"T-{t['tower_number']}" for t in listed],
        )

    def test_export_cannot_escape_the_viewers_scope(self):
        """An employee with no assignment oversees nothing, so both exports come
        back as valid-but-empty reports rather than leaking another
        jurisdiction's records."""
        client = self._client('NONE1')
        history = client.get(self.HISTORY)
        self.assertEqual(history.status_code, 200)
        self.assertEqual(self._table(history), [])
        tickets = client.get(self.TICKETS)
        self.assertEqual(tickets.status_code, 200)
        self.assertEqual(self._table(tickets), [])

    def test_a_tower_filter_outside_scope_yields_nothing(self):
        """?tower only ever narrows — pointing it at someone else's tower cannot
        widen the report to include it."""
        other_sub = Subdivision.objects.create(name='Y SD', zone='YZ')
        other_line = Line.objects.create(
            source_layer='t', arcgis_object_id='YL1', name='Y LINE', voltage='132kV',
            subdivision=other_sub, is_active=True,
        )
        other_tower = Tower.objects.create(
            source_layer='t', arcgis_object_id='YT1', tower_number='99', voltage='132kV',
            line=other_line, line_name='Y LINE', is_active=True,
        )
        resp = self._client('AEE1').get(f'{self.HISTORY}?tower={other_tower.id}')
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(self._table(resp), [])

    def test_unknown_format_is_a_400_not_a_wrong_file(self):
        for url in (self.HISTORY, self.TICKETS):
            resp = self._client('AEE1').get(url + '?format=docx')
            self.assertEqual(resp.status_code, 400, url)
            self.assertIn('format', resp.json())

    def test_exports_require_a_session(self):
        for url in (self.HISTORY, self.TICKETS):
            self.assertIn(Client().get(url).status_code, (401, 403), url)
